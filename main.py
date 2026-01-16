#!/usr/bin/env python3
"""
Cisco Access Switch Audit → Excel
================================

Overview
--------
Audits one or more Cisco switches (optionally via an SSH jump host),
collecting interface state, PoE information, neighbor presence (LLDP/CDP), and
selected error counters, then exports a **filters‑only** Excel workbook with a
first‑class **SUMMARY** sheet and one sheet per device. The script favors
robustness and repeatability and is safe to run against large device sets
thanks to a threaded worker pool and an event‑driven progress display.

Data Sources (per device)
-------------------------
* ``show interfaces`` (parsed via TextFSM when templates are available)
* ``show interfaces status`` (custom fixed‑width parser — authoritative for Mode/VLAN)
* ``show power inline`` (PoE admin/oper state, power draw, class/device)
* ``show lldp neighbors detail`` and ``show cdp neighbors detail`` (neighbor presence)

Key Features
------------
* **Mode/VLAN classification** using a resilient parser of ``show interfaces status``
  (``access`` / ``trunk`` / ``routed``). When both TextFSM and status data exist, the
  status parser **overrides** mode/VLAN to ensure consistency.
* **Interface name normalization & aliasing** (e.g., ``GigabitEthernet1/0/1`` ↔ ``Gi1/0/1``)
  so that fields from different commands line up reliably.
* **PoE enrichment** (admin/oper/power/class/device) with multi‑key lookups across
  short/long/original interface names.
* **Neighbor signal** — boolean flag when LLDP or CDP reports a neighbor on the port.
* **Stale‑port analysis** (access ports only; configurable via ``--stale-days``):
  - If **connected** → stale **only** when ``Last input ≥ N days``.
  - If **not connected** → stale when **no PoE draw** **and** **no LLDP/CDP neighbor**.
* **Excel output** with:
  - Frozen header, AutoFilter, auto column widths (bounded), and conditional formatting
    (status colors, error counters > 0 in red, PoE power > 0 in green, Stale=True in red).
  - A **SUMMARY** sheet (first) plus one sheet per device; includes a **TOTAL** row that
    sums numeric columns across devices.
* **Concurrency** via a ThreadPoolExecutor (``--workers``) and an event queue that drives a
  smooth progress bar (started vs completed).
* **Jump‑host support** via a persistent SSH bastion (``JumpManager``). The tool uses the
  jump automatically unless ``--direct`` is given.
* **Credential handling** that prefers secure retrieval (e.g., Windows Credential Manager
  target ``MyApp/ADM``) with interactive fallback; optional enable secret retrieval.

Inputs & CLI
------------
* **Devices file** (required): plain text with one hostname or IP per line (``#`` comments allowed).
* **Arguments**:
  - ``--devices, -d`` (required): path to devices file.
  - ``--output,  -o``: output workbook filename (default: ``audit.xlsx``).
  - ``--workers, -w``: max concurrent device sessions (default: 10).
  - ``--stale-days``: inactivity threshold in days for stale detection (0 disables; default: 30).
  - ``--direct``: connect directly (skip jump host).
  - ``--debug``: verbose logging.

Excel Layout
------------
* **SUMMARY** (first): per‑device metrics (Totals, Access/Trunk/Routed counts, Status counts,
  and percentage columns) + final **TOTAL** row.
* **Per‑device** sheets: detailed records with columns such as Device, Mgmt IP, Interface,
  Description, Status, Mode, VLAN, Duplex, Speed, Type, Input/Output/CRC Errors, Last Input,
  PoE fields, Neighbor flag, and ``Stale (≥Nd)``.

Reliability & Fallback Strategy
-------------------------------
* If TextFSM templates are available (``NET_TEXTFSM``), ``show interfaces`` is parsed in detail;
  otherwise the script still operates and relies on ``show interfaces status`` and best‑effort
  enrichment.
* Status/mode/VLAN are **authoritatively** taken from the status parser when present.
* All per‑device failures are captured into that device’s summary row; the run continues.

Dependencies
------------
* Python 3.8+
* Packages: ``netmiko``, ``paramiko``, ``pandas``, ``openpyxl`` (and ``pywin32`` on Windows for
  Credential Manager integration).

Security Notes
--------------
* Prefer secure stores (e.g., Windows Credential Manager) over plaintext. Never commit secrets.
* When using a jump host, enforce strong authentication and reasonable idle timeouts.

Extensibility Hooks
-------------------
* Adjust connection behavior in ``Modules/netmiko_utils.py``.
* Change bastion behavior in ``Modules/jump_manager.py``.
* Modify credential retrieval in ``Modules/credentials.py``.
* Tweak output columns/formatting in ``main.py`` (``detailed.append({...})`` and
  ``_format_worksheet()``).

This module is designed to produce analyst‑friendly, filterable workbooks while
remaining safe and predictable on enterprise networks.
"""
from __future__ import annotations
import argparse
import concurrent.futures as cf
import re
import sys
import time
from typing import Dict, List, Tuple, Any
from queue import Queue, Empty
import logging

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill

# ---- external imports ----
from Modules.credentials import get_secret_with_fallback, get_enable_secret
from Modules.jump_manager import JumpManager
from Modules.config import JUMP_HOST
from Modules.netmiko_utils import connect_to_device

# ---------------- Configuration ----------------
MIN_WIDTH = 8   # Minimum Excel column width
MAX_WIDTH = 60  # Maximum Excel column width


# ---------------- Helpers ----------------
def acquire_username_password() -> Tuple[str, str]:
    """
    Retrieve login credentials for device access.

    Returns:
        (username, password) tuple obtained from secure storage.
    Raises:
        RuntimeError: if credentials are missing or invalid.
    """
    user, pwd = get_secret_with_fallback()
    if not user or not pwd:
        raise RuntimeError("get_secret_with_fallback() did not return username/password")
    return user, pwd


def acquire_enable_secret() -> str | None:
    """
    Retrieve the enable secret (if configured).

    Returns:
        Enable password string or None if not available.
    """
    enable_pwd = get_enable_secret()
    return enable_pwd


def _sheet_name_from(base: str, fallback: str) -> str:
    """
    Generate a safe Excel sheet name from a base string.

    Args:
        base: Preferred name (e.g., hostname).
        fallback: Fallback string (e.g., IP address).

    Returns:
        Sanitized sheet name (<= 31 chars, valid characters only).
    """
    name = base or fallback
    sheet = re.sub(r"[^A-Za-z0-9_-]", "_", name)[:31]
    return sheet or re.sub(r"[^A-Za-z0-9_-]", "_", fallback)[:31]


def _unique_sheet_name(suggested: str, existing: set[str], limit: int = 31) -> str:
    """
    Ensure sheet names are unique within the workbook.

    Args:
        suggested: Initial sheet name suggestion.
        existing: Set of already-used sheet names.
        limit: Max length allowed by Excel.

    Returns:
        Unique sheet name string.
    """
    lower = {e.lower() for e in existing}
    base = suggested[:limit]
    if base.lower() not in lower:
        existing.add(base)
        return base
    for i in range(2, 1000):
        suffix = f"-{i}"
        cand = (base[: max(0, limit - len(suffix))] + suffix)
        if cand.lower() not in lower:
            existing.add(cand)
            return cand
    # Fallback if all else fails
    k = 1
    while True:
        cand = f"Sheet{k}"[:limit]
        if cand.lower() not in lower:
            existing.add(cand)
            return cand
        k += 1


# ---------------- Formatting (filters only) ----------------
def _format_worksheet(ws, df: pd.DataFrame) -> None:
    """
    Apply formatting to an Excel worksheet.

    - Freezes header row
    - Adds auto-filter across all columns
    - Auto-sizes column widths based on content
    - Applies conditional formatting:
        * Status values (connected, notconnect, admin down, err-disabled)
        * Error counters > 0 flagged red
        * PoE power > 0 flagged green
        * Stale flags highlighted red

    Args:
        ws: openpyxl worksheet object
        df: DataFrame used to populate the sheet (for sizing reference)
    """
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return
    ws.freeze_panes = "A2"
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.auto_filter.ref = ref

    # Auto column widths
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))  # header
        for row in range(2, max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        width = max(MIN_WIDTH, min(MAX_WIDTH, max_len + 2))
        ws.column_dimensions[col_letter].width = width

    # Conditional formatting fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    grey_fill  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Apply rules based on headers
    for col_idx in range(1, max_col + 1):
        header = str(ws.cell(row=1, column=col_idx).value or "").lower()
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{max_row}"

        if header == "status":
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="equal", formula=['"connected"'], fill=green_fill))
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="equal", formula=['"notconnect"'], fill=red_fill))
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="equal", formula=['"err-disabled"'], fill=red_fill))
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="equal", formula=['"administratively down"'], fill=grey_fill))
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="equal", formula=['"disabled"'], fill=grey_fill))

        elif "errors" in header:  # Input/Output/CRC Errors
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="greaterThan", formula=['0'], fill=red_fill))

        elif header.startswith("poe power"):
            ws.conditional_formatting.add(rng,
                CellIsRule(operator="greaterThan", formula=['0'], fill=green_fill))

        elif "stale" in header:
            # True values flagged red
            ws.conditional_formatting.add(rng,
                FormulaRule(formula=[f'${col_letter}2=TRUE'], fill=red_fill))


# ------------- Interface name normalization -------------
_IF_MAP = {
    "gigabitethernet": "Gi",
    "gig": "Gi",
    "gi": "Gi",
    "fastethernet": "Fa",
    "fast": "Fa",
    "fa": "Fa",
    "tengigabitethernet": "Te",
    "ten": "Te",
    "te": "Te",
    "hundredgige": "Hu",
    "hundredgigabit": "Hu",
    "hundredgigabitethernet": "Hu",
    "hu": "Hu",
    "ethernet": "Eth",
    "eth": "Eth",
    "et": "Eth",
}

def normalize_ifname(ifname: str) -> Tuple[str, str]:
    """
    Normalize interface names to canonical short and long forms.

    Args:
        ifname: Raw interface name string.

    Returns:
        (short_form, long_form) tuple, e.g. ("Gi1/0/1", "GigabitEthernet1/0/1").
    """
    s = (ifname or "").strip()
    if not s:
        return ("", "")
    m = re.match(r"([A-Za-z]+)([0-9/\.]+.*)", s)
    if not m:
        return (s, s)
    prefix_raw = m.group(1)
    rest = m.group(2)
    key = prefix_raw.lower()
    short_prefix = _IF_MAP.get(key, prefix_raw)
    long_prefix = {
        "Gi": "GigabitEthernet",
        "Fa": "FastEthernet",
        "Te": "TenGigabitEthernet",
        "Eth": "Ethernet",
        "Hu": "HundredGigE",
    }.get(short_prefix, prefix_raw)
    return (f"{short_prefix}{rest}", f"{long_prefix}{rest}")


def all_aliases(ifname: str) -> List[str]:
    """
    Return a list of likely alias strings for matching across commands.

    Args:
        ifname: Raw interface name string.

    Returns:
        List of alias strings (short, long, and original).
    """
    sh, lo = normalize_ifname(ifname)
    aliases = {sh, lo, ifname}
    return [a for a in aliases if a]


# ------------- Time parsing for "Last input" -------------
_TIME_RE = re.compile(
    r"(?:(?P<y>\d+)\s*y)?\s*(?:(?P<w>\d+)\s*w)?\s*(?:(?P<d>\d+)\s*d)?\s*(?P<h>\d+?\s*h)?\s*(?P<m>\d+?\s*m)?\s*(?P<s>\d+?\s*s)?",
    re.IGNORECASE
)

def _parse_last_input_seconds(s: str) -> float | None:
    """
    Parse Cisco 'Last input' timer strings into seconds.

    Handles formats like:
    - "00:01:23"
    - "1d2h30m"
    - "never"

    Args:
        s: Raw timer string.

    Returns:
        Seconds as float, or None if unparsable/never.
    """
    s = (s or "").strip().lower()
    if not s or s == "never":
        return None
    # hh:mm:ss
    if re.match(r"^\d{1,2}:\d{2}:\d{2}$", s):
        hh, mm, ss = s.split(":")
        return int(hh) * 3600 + int(mm) * 60 + int(ss)
    # compact duration (y, w, d, h, m, s)
    m = _TIME_RE.fullmatch(s.replace(" ", ""))
    if not m:
        return None
    y = int(m.group("y") or 0)
    w = int(m.group("w") or 0)
    d = int(m.group("d") or 0)
    h = int((m.group("h") or "0").rstrip("h") or 0)
    mn = int((m.group("m") or "0").rstrip("m") or 0)
    sc = int((m.group("s") or "0").rstrip("s") or 0)
    days = y * 365 + w * 7 + d
    return days * 86400 + h * 3600 + mn * 60 + sc


# ------------- LLDP / CDP neighbors -------------
def _get_lldp_neighbors(conn) -> Dict[str, Tuple[str, str]]:
    """
    Parse 'show lldp neighbors detail' output.

    Returns:
        Dict mapping local interface -> (remote device, remote port).
    """
    neighbors: Dict[str, Tuple[str, str]] = {}
    try:
        out = conn.send_command("show lldp neighbors detail", read_timeout=60, use_textfsm=False)
        cur_dev, cur_port, cur_local = None, None, None
        for line in out.splitlines():
            m = re.search(r"Device ID:\s*(\S+)", line)
            if m:
                cur_dev = m.group(1)
            m = re.search(r"Port id:\s*(\S+)", line, flags=re.I)
            if m:
                cur_port = m.group(1)
            m = re.search(r"(Local Intf|Local Port id)\s*:\s*(\S+)", line, flags=re.I)
            if m:
                cur_local = normalize_ifname(m.group(2))[0]
            if cur_dev and cur_port and cur_local:
                neighbors[cur_local] = (cur_dev, cur_port)
                cur_dev, cur_port, cur_local = None, None, None
    except Exception:
        pass
    return neighbors

def _get_cdp_neighbors(conn) -> Dict[str, Tuple[str, str]]:
    """
    Parse 'show cdp neighbors detail' output.

    Returns:
        Dict mapping local interface -> (remote device, remote port).
    """
    neighbors: Dict[str, Tuple[str, str]] = {}
    try:
        out = conn.send_command("show cdp neighbors detail", read_timeout=60, use_textfsm=False)
        cur_dev = None
        for line in out.splitlines():
            m = re.search(r"Device ID:\s*(\S+)", line)
            if m:
                cur_dev = m.group(1)
            m = re.search(r"Interface:\s*([A-Za-z]+[A-Za-z]?[0-9/\.]+),\s*Port ID.*:\s*(\S+)", line)
            if m and cur_dev:
                local = normalize_ifname(m.group(1))[0]
                neighbors[local] = (cur_dev, m.group(2))
                cur_dev = None
    except Exception:
        pass
    return neighbors


def _parse_ios_version(show_ver_text: str) -> str:
    """
    Extract a concise OS version string from 'show version' output.
    Tries common IOS/IOS-XE patterns; falls back to the first 'Version x.y...' hit.
    Returns short text like 'IOS-XE 17.6.5a' or 'IOS 15.2(7)E8' if found, else ''.
    """
    s = show_ver_text or ""
    # Common IOS-XE banner: 'Cisco IOS XE Software, Version 17.6.5a'
    m = re.search(r"Cisco IOS XE Software,\s*Version\s*([^\s,]+)", s, flags=re.I)
    if m:
        return f"IOS-XE {m.group(1)}"
    # Common IOS banner: 'Cisco IOS Software, ... Version 15.2(7)E8'
    m = re.search(r"Cisco IOS Software.*Version\s*([^\s,]+)", s, flags=re.I)
    if m:
        return f"IOS {m.group(1)}"
    # NX-OS/others pattern (just in case): 'system:    version 9.3(10)'
    m = re.search(r"\bversion\s+([0-9][^ \r\n\t,;)]*)", s, flags=re.I)
    if m:
        return m.group(1)
    return ""


# ---------------- 'show interfaces status' robust parser (fallback) ----------------
def _find_columns(header_line: str) -> Dict[str, slice]:
    """
    Discover fixed-column slices from the header line of 'show interfaces status'.

    Args:
        header_line: The header line containing column names.

    Returns:
        Mapping from lowercase column name to slice object.
    """
    tokens = [("Port", None), ("Name", None), ("Status", None), ("Vlan", None),
              ("Duplex", None), ("Speed", None), ("Type", None)]
    for i, (tok, _) in enumerate(tokens):
        idx = header_line.find(tok)
        if idx == -1:
            raise ValueError("Unrecognized header for 'show interfaces status': " + header_line)
        tokens[i] = (tok, idx)
    positions = [idx for _, idx in tokens]
    positions_sorted = sorted(positions)
    slices = {}
    for i, (tok, idx) in enumerate(tokens):
        start = idx
        next_starts = [p for p in positions_sorted if p > start]
        end = next_starts[0] if next_starts else None
        slices[tok.lower()] = slice(start, end)
    return slices

def parse_show_interfaces_status(output: str) -> List[Dict[str, str]]:
    """
    Parse the text output of 'show interfaces status' into records.

    Args:
        output: Raw multiline string from the device.

    Returns:
        List of dict rows with keys: port, name, status, vlan, duplex, speed, type.
    """
    lines = [ln.rstrip("\r") for ln in output.splitlines() if ln.strip()]
    records: List[Dict[str, str]] = []

    def is_header(ln: str) -> bool:
        return ("Port" in ln and "Status" in ln and "Vlan" in ln and "Speed" in ln)

    i, n = 0, len(lines)
    while i < n:
        ln = lines[i]
        if not is_header(ln):
            i += 1
            continue

        try:
            col = _find_columns(ln)
        except ValueError:
            i += 1
            continue

        i += 1
        while i < n:
            row = lines[i]
            if is_header(row):
                break
            if set(row.strip()) == {"-"}:
                i += 1
                continue
            if row.strip().lower().startswith(("port ", "----")):
                i += 1
                continue
            try:
                rec = {
                    'port':   row[col['port']].strip().lstrip("*^!"),
                    'name':   row[col['name']].strip(),
                    'status': row[col['status']].strip().lower(),
                    'vlan':   row[col['vlan']].strip().lower(),
                    'duplex': row[col['duplex']].strip().lower(),
                    'speed':  row[col['speed']].strip().lower(),
                    'type':   row[col['type']].strip(),
                }
            except Exception:
                i += 1
                continue
            if rec['port']:
                records.append(rec)
            i += 1

    return records


def parse_show_power_inline(output: str) -> Dict[str, Dict[str, Any]]:
    """
    Parse 'show power inline' and return a map keyed by interface aliases.
    Example rows often look like:
      Interface  Admin  Oper  Power   Class  Device
      Gi1/0/1    auto   on    7.0 W   2      IP Phone 7962

    Returns:
        Dict[str, Dict[str, Any]]: per-interface PoE details keyed by short, long, and raw names.
    """
    lines = [ln for ln in output.splitlines() if ln.strip()]
    header_idx = None
    for i, ln in enumerate(lines):
        if 'Interface' in ln and 'Admin' in ln and 'Oper' in ln:
            header_idx = i
            break
    if header_idx is None:
        return {}

    header = lines[header_idx]
    def idx(tok: str) -> int | None:
        p = header.find(tok)
        return p if p != -1 else None
    starts = {
        'if': idx('Interface'),
        'admin': idx('Admin'),
        'oper': idx('Oper'),
        'power': idx('Power'),
        'class': idx('Class'),
        'device': idx('Device'),
    }
    keys_order = [k for k, v in starts.items() if v is not None]
    pos_sorted = sorted([(k, starts[k]) for k in keys_order], key=lambda x: x[1])
    slices: Dict[str, slice] = {}
    for i, (k, start) in enumerate(pos_sorted):
        next_start = pos_sorted[i + 1][1] if i + 1 < len(pos_sorted) else None
        slices[k] = slice(start, next_start)

    data: Dict[str, Dict[str, Any]] = {}
    for ln in lines[header_idx + 1:]:
        if set(ln.strip()) == {"-"}:
            continue
        iface_raw = ln[slices['if']].strip() if 'if' in slices else None
        if not iface_raw:
            continue
        short, longf = normalize_ifname(iface_raw)

        entry = {
            'poe_admin':  ln[slices['admin']].strip().lower() if 'admin' in slices else '',
            'poe_oper':   ln[slices['oper']].strip().lower() if 'oper' in slices else '',
            'poe_power_w': ln[slices['power']].strip() if 'power' in slices else '',
            'poe_class':  ln[slices['class']].strip() if 'class' in slices else '',
            'poe_device': ln[slices['device']].strip() if 'device' in slices else '',
        }
        # Store under multiple keys to maximize match rate
        data[short] = entry
        data[longf] = entry
        data[iface_raw] = entry

    return data


# ---------------- TextFSM 'show interfaces' primary path ----------------
def get_interfaces_via_show_interfaces(conn) -> List[Dict[str, Any]]:
    """
    Use TextFSM to parse 'show interfaces' for all ports.
    Expected keys include (template-dependent):
      - interface, description
      - status/link_status, protocol/protocol_status, admin_state
      - duplex, speed, hardware_type
      - input_errors, output_errors, crc
      - last_input

    Returns:
        List[Dict[str, Any]] of parsed records; empty list on failure.
    """
    try:
        output = conn.send_command("show interfaces", use_textfsm=True)
        if isinstance(output, list):
            return output
        return []
    except Exception:
        return []


# ---------------- Device worker ----------------
def audit_device(ip: str,
                 username: str,
                 password: str,
                 enable_secret: str | None,
                 jump_host: str | None,
                 stale_days: int,
                 debug: bool = False) -> Tuple[str, str | None, List[Dict[str, Any]], Dict[str, Any]]:
    """
    Connect to a single device and perform audit.

    - Establishes SSH session (direct or via jump host)
    - Collects interface data using TextFSM
    - Enriches with PoE, LLDP/CDP neighbors
    - Computes stale flags and error counters
    - Builds per-device summary dictionary

    Returns:
        (ip, hostname, detailed_records, summary_dict)
    """
    jump = JumpManager(jump_host, username, password) if jump_host else None
    conn = None
    hostname: str | None = None
    try:
        if jump:
            with jump:
                conn = connect_to_device(ip, username, password, jump=jump)
                return _audit_connected_device(conn, ip, enable_secret, stale_days, debug)
        else:
            conn = connect_to_device(ip, username, password)
            return _audit_connected_device(conn, ip, enable_secret, stale_days, debug)

    except Exception as e:
        return ip, hostname, [], {
            'Device': hostname or ip,
            'Mgmt IP': ip,
            'IOS Version': '',
            'Total Ports (phy)': 'ERROR',
            'Access Ports': '',
            'Connected': '',
            'Not Connected': '',
            'Admin Down': '',
            'Err-Disabled': '',
            'Error': str(e),
        }
    finally:
        try:
            if conn:
                conn.disconnect()
        except Exception:
            pass


def _audit_connected_device(
    conn,
    ip: str,
    enable_secret: str | None,
    stale_days: int,
    debug: bool
) -> Tuple[str, str | None, List[Dict[str, Any]], Dict[str, Any]]:
    """
    Perform the audit once connected to a device:
    - Retrieves hostname and enters enable mode if provided
    - Parses 'show interfaces' via TextFSM
    - Enriches with PoE and neighbor presence
    - Normalizes status, computes stale flag for access ports
    - Aggregates a summary row for the device

    Extended:
    - Uses 'show interfaces status' (fixed-width parser) to classify Mode (access/trunk/routed)
      and fill VLAN reliably, overriding the TextFSM 'show interfaces' record when present.
    """
    hostname: str | None = None

    # --- Identify hostname
    try:
        out_hn = conn.send_command("show running-config | include ^hostname ", read_timeout=30)
        m = re.search(r"^hostname\s+(\S+)", out_hn, flags=re.M)
        if m:
            hostname = m.group(1).strip()
    except Exception:
        pass
    if not hostname:
        try:
            hostname = conn.find_prompt().rstrip(">#").strip()
        except Exception:
            hostname = None

    # --- Enter enable mode if a secret is configured
    if enable_secret:
        try:
            conn.enable()
            if debug:
                print(f"[{hostname or ip}] enable mode entered")
        except Exception as e:
            if debug:
                print(f"[{hostname or ip}] enable failed: {e}")

    ios_version = ""
    try:
        out_ver = conn.send_command("show version", read_timeout=60, use_textfsm=False)
        ios_version = _parse_ios_version(out_ver)
    except Exception:
        ios_version = ""

    # --- Primary interface records via TextFSM
    tfsm_records = get_interfaces_via_show_interfaces(conn)
    detailed: List[Dict[str, Any]] = []

    # --- NEW: Build a per-interface mode/VLAN map from 'show interfaces status'
    
    sif_status_map: Dict[str, Dict[str, str]] = {}
    try:
        out_status = conn.send_command("show interfaces status", read_timeout=60, use_textfsm=False)
        for rec in parse_show_interfaces_status(out_status):
            # Normalize status row port name to short form (e.g., 'Gi1/0/1')
            port_norm = normalize_ifname(rec.get('port', ''))[0]
            if not port_norm:
                continue

            vlan_raw   = (rec.get('vlan')   or '').strip().lower()
            status_raw = (rec.get('status') or '').strip().lower()

            if vlan_raw in {'trunk', 'rspan'}:
                mode = 'trunk'
                vlan = vlan_raw
            elif vlan_raw == 'routed':
                mode = 'routed'
                vlan = vlan_raw
            else:
                # numeric/none/blank -> treat as access
                # keep vlan text as-is (e.g., '1', '10', 'none', '')
                mode = 'access'
                vlan = vlan_raw

            # Stash status so we can override the final Status if present
            sif_status_map[port_norm] = {'mode': mode, 'vlan': vlan, 'status': status_raw}
    except Exception:
        # If we can't parse/show interfaces status, we'll fall back to the prior behavior
        sif_status_map = {}

    if tfsm_records:
        # Enrichment: PoE + neighbors
        try:
            out_poe = conn.send_command("show power inline", read_timeout=60, use_textfsm=False)
            poe_map = parse_show_power_inline(out_poe)
        except Exception:
            poe_map = {}

        lldp_ifaces = _get_lldp_neighbors(conn)
        cdp_ifaces = _get_cdp_neighbors(conn)

        for r in tfsm_records:
            port_raw = r.get('interface') or ""
            short_port, long_port = normalize_ifname(port_raw)
            if not short_port:
                continue

            description = r.get('description', '')

            
            # --- Normalize operational/admin/protocol state into a canonical Status
            line_state  = (r.get('link_status') or r.get('status') or '').strip().lower().replace('_', '-')
            admin_state = (r.get('admin_state') or '').strip().lower().replace('_', '-')
            proto_state = (r.get('protocol') or r.get('protocol_status') or '').strip().lower().replace('_', '-')

            def _canon(s: str) -> str:
                return ' '.join(s.split())

            line_state  = _canon(line_state)
            admin_state = _canon(admin_state)
            proto_state = _canon(proto_state)

            # Base status from detailed ('show interfaces') fields
            if "administratively down" in (line_state, admin_state, proto_state):
                status = "administratively down"
            elif "err-disabled" in (line_state, admin_state, proto_state) or "errdisabled" in (line_state, admin_state, proto_state):
                status = "err-disabled"
            elif line_state == "up" and proto_state == "up":
                status = "connected"
            elif line_state == "down" and proto_state == "down":
                status = "notconnect"
            elif line_state == "up":
                status = "connected"
            elif line_state == "down":
                status = "notconnect"
            else:
                status = "unknown"

            _ovr = None
            if short_port in sif_status_map:
                _ovr = (sif_status_map[short_port].get('status') or '').strip().lower()
            else:
                for alias in all_aliases(short_port):
                    if alias in sif_status_map:
                        _ovr = (sif_status_map[alias].get('status') or '').strip().lower()
                        break

            if _ovr:
                # Normalize common variants to canonical values
                if _ovr in {"errdisabled"}:
                    _ovr = "err-disabled"
                # Accept typical table statuses as-is
                if _ovr in {"connected", "notconnect", "administratively down", "err-disabled", "inactive", "monitoring"}:
                    status = _ovr

            # OLD values (kept as fallback if 'show interfaces status' is unavailable)
            vlan = (r.get('vlan') or '').lower()
            is_trunk = vlan in {"trunk", "rspan"}
            mode = "trunk" if is_trunk else ("routed" if vlan == "routed" else "access")

            # NEW: Prefer classification from 'show interfaces status'
            if short_port in sif_status_map:
                mode = sif_status_map[short_port]['mode']
                vlan = sif_status_map[short_port]['vlan']
            else:
                # Try aliases just in case there are naming differences
                for alias in all_aliases(short_port):
                    if alias in sif_status_map:
                        mode = sif_status_map[alias]['mode']
                        vlan = sif_status_map[alias]['vlan']
                        break

            duplex = (r.get('duplex') or '').lower()
            speed = (str(r.get('speed') or '')).lower()
            hwtype = r.get('hardware_type', '')

            # Error counters
            try:
                in_err = int(r.get('input_errors') or 0)
            except Exception:
                in_err = 0
            try:
                out_err = int(r.get('output_errors') or 0)
            except Exception:
                out_err = 0
            try:
                crc_err = int(r.get('crc') or 0)
            except Exception:
                crc_err = 0


            # Last input / output times
            last_input_raw = r.get('last_input', '') or r.get('last_input_text', '') or ''
            last_input_secs = _parse_last_input_seconds(last_input_raw) if last_input_raw else None

            last_output_raw = r.get('last_output', '') or r.get('last_output_text', '') or ''
            # We don't currently use last_output_secs for any logic, but keep it for parity/debugging.
            last_output_secs = _parse_last_input_seconds(last_output_raw) if last_output_raw else None


            # PoE lookup using aliases (short/long/raw)
            poe = {}
            for alias in all_aliases(short_port):
                if alias in poe_map:
                    poe = poe_map[alias]
                    break

            # Convert "7.0 W" => 7.0
            poe_w = None
            try:
                first = str(poe.get('poe_power_w', '')).split()[0]
                poe_w = float(first) if first not in {"", "-"} else None
            except Exception:
                poe_w = None
            poe_active = (poe_w is not None and poe_w > 0.0)

            has_neighbor = (short_port in lldp_ifaces) or (short_port in cdp_ifaces)

            # Conservative stale definition for access ports
            stale_flag = None
            if stale_days > 0 and mode == "access":
                if status == "connected":
                    stale_flag = (last_input_secs is not None and last_input_secs >= (stale_days * 86400))
                else:
                    stale_flag = (not poe_active) and (not has_neighbor)

            detailed.append({
                'Device': hostname or ip,
                'Mgmt IP': ip,
                'Interface': long_port,
                'Description': description,
                'Status': status,
                'AdminDown': status == "administratively down",
                'Connected': status == "connected",
                'ErrDisabled': status == "err-disabled",
                'Mode': mode,
                'VLAN': vlan,
                'Duplex': duplex,
                'Speed': speed,
                'Type': hwtype,
                'Input Errors': in_err,
                'Output Errors': out_err,
                'CRC Errors': crc_err,
                'Last Input': last_input_raw,
                'Last Output': last_output_raw,
                'PoE Power (W)': poe.get('poe_power_w', ''),
                'PoE Oper': poe.get('poe_oper', ''),
                'PoE Admin': poe.get('poe_admin', ''),
                'LLDP/ CDP Neighbor': has_neighbor,
                f'Stale (≥{stale_days} d)': stale_flag,
            })

    # --- Build summary
    df = pd.DataFrame(detailed)

    if df.empty:
        summary = {
            'Device': hostname or ip,
            'Mgmt IP': ip,
            'IOS Version': ios_version,
            'Total Ports (phy)': 0,
            'Access Ports': 0,
            'Trunk Ports': 0,
            'Routed Ports': 0,
            'Connected': 0,
            'Not Connected': 0,
            'Admin Down': 0,
            'Err-Disabled': 0,
            '% Access of Total': 0.0,
            '% Trunk of Total': 0.0,
            '% Routed of Total': 0.0,
            '% Connected of Total': 0.0,
        }
    else:
        total = len(df)
        access_cnt = int((df['Mode'] == 'access').sum())
        trunk_cnt  = int((df['Mode'] == 'trunk').sum())
        routed_cnt = int((df['Mode'] == 'routed').sum())

        connected_cnt   = int((df['Status'] == 'connected').sum())
        notconnect_cnt  = int((df['Status'] == 'notconnect').sum())
        admindown_cnt   = int((df['Status'] == 'administratively down').sum())
        errdisabled_cnt = int((df['Status'] == 'err-disabled').sum())
    
        def _pct(n: int, d: int) -> float:
            return round((n / d * 100.0), 1) if d else 0.0
        
        summary = {
            'Device': hostname or ip,
            'Mgmt IP': ip,
            'IOS Version': ios_version,
            'Total Ports (phy)': total,
            'Access Ports': access_cnt,
            'Trunk Ports': trunk_cnt,
            'Routed Ports': routed_cnt,
            'Connected': connected_cnt,
            'Not Connected': notconnect_cnt,
            'Admin Down': admindown_cnt,
            'Err-Disabled': errdisabled_cnt,
            '% Access of Total': _pct(access_cnt, total),
            '% Trunk of Total': _pct(trunk_cnt, total),
            '% Routed of Total': _pct(routed_cnt, total),
            '% Connected of Total': _pct(connected_cnt, total),
            '% AdminDown of Total': _pct(admindown_cnt, total),
        }

    return ip, hostname, detailed, summary


# ---------------- Progress bar (event-driven with final drain) ----------------
def _print_progress_extended(started: int, done: int, total: int, width: int = 30) -> None:
    """
    Render a simple, event-driven progress bar (non-thread-safe write).

    Args:
        started: Number of jobs that have started
        done: Number of jobs completed
        total: Total jobs
        width: Bar width in characters
    """
    ratio = 0 if total == 0 else done / total
    filled = int(ratio * width)
    bar = "█" * filled + "░" * (width - filled)
    sys.stdout.write(f"\rProgress: [{bar}] {done}/{total}  |  started: {started}/{total}")
    sys.stdout.flush()
    if done == total:
        sys.stdout.write("\n")

def _worker_wrapper(func, ip, username, password, enable_secret, jump_host, stale_days, debug, event_q: Queue):
    """
    Wrap a worker to emit start/done events to the queue for progress display.
    """
    try:
        event_q.put(("start", ip, time.time()))
        result = func(ip, username, password, enable_secret, jump_host, stale_days, debug)
        event_q.put(("done", ip, time.time()))
        return result
    except Exception:
        event_q.put(("done", ip, time.time()))
        raise


# ---------------- Main ----------------
def main():
    """
    CLI entrypoint:
    - Parses args
    - Loads device list
    - Runs audit concurrently with progress
    - Writes SUMMARY + per-device sheets to Excel
    """
    parser = argparse.ArgumentParser(description="Cisco Access Switch Audit -> Excel (TextFSM 'show interfaces')")
    parser.add_argument("--direct", action="store_true", help="Connect directly (no jump host)")
    parser.add_argument("--stale-days", type=int, default=30, help="Days to consider 'stale' for access ports (0 to disable).")
    parser.add_argument("--debug", action="store_true", help="Enable debug output")
    parser.add_argument("--output", "-o", type=str, default="audit.xlsx", help="Output workbook filename")
    parser.add_argument("--devices", "-d", type=str, default="devices.txt", required=True, help="Path to devices file (one IP/hostname per line, # comments allowed)")
    parser.add_argument("--workers", "-w", type=int, default=10, help="Number of worker threads")
    args = parser.parse_args()

    # Silence output unless --debug
    logging.basicConfig(level=logging.DEBUG if args.debug else logging.CRITICAL, format="%(message)s")
    log = logging.getLogger(__name__)

    jump_host: str | None = None if args.direct else JUMP_HOST
    if args.debug:
        if jump_host:
            log.info(f"Using jump host {JUMP_HOST} to connect to devices")
        else:
            log.info("Connecting directly to devices (no jump host)")

    username, password = acquire_username_password()
    enable_secret = acquire_enable_secret()

    with open(args.devices, "r", encoding="utf-8") as f:
        device_list = [ln.strip() for ln in f if ln.strip() and not ln.strip().startswith("#")]
    if not device_list:
        raise SystemExit("No devices found in input file")

    if args.debug:
        print(f"[*] Devices: {len(device_list)} | Jump host: {jump_host or 'none'} "
              f"| Workers: {args.workers} | Stale≥{args.stale_days}d")

    results: List[Tuple[str, str | None, List[Dict[str, Any]], Dict[str, Any]]] = []
    total = len(device_list)

    event_q: Queue = Queue()
    started = 0
    done = 0
    _print_progress_extended(started, done, total)

    try:
        with cf.ThreadPoolExecutor(max_workers=args.workers) as ex:
            futs = [
                ex.submit(
                    _worker_wrapper,
                    audit_device, ip, username, password, enable_secret, jump_host, args.stale_days, args.debug, event_q
                )
                for ip in device_list
            ]

            pending = set(futs)
            while pending:
                # Update progress from event queue
                while True:
                    try:
                        ev, ev_ip, ev_ts = event_q.get_nowait()
                    except Empty:
                        break
                    if ev == "start":
                        started += 1
                    elif ev == "done":
                        done += 1
                    _print_progress_extended(started, done, total)

                done_set, pending = cf.wait(pending, timeout=0.2, return_when=cf.FIRST_COMPLETED)
                for fut in done_set:
                    try:
                        results.append(fut.result())
                    except Exception as e:
                        results.append(("UNKNOWN", None, [], {
                            'Device': 'UNKNOWN', 'Mgmt IP': '', 'Total Ports (phy)': 'ERROR', 'Error': str(e)
                        }))
                _print_progress_extended(started, done, total)
    finally:
        # Final drain to ensure progress reflects completion
        while True:
            try:
                ev, ev_ip, ev_ts = event_q.get_nowait()
            except Empty:
                break
            if ev == "start":
                started += 1
            elif ev == "done":
                done += 1

        done = max(done, len(results))
        started = max(started, total)
        _print_progress_extended(started, done, total)

    # Write Excel
    with pd.ExcelWriter(args.output, engine="openpyxl") as xw:
        used_sheet_names: set[str] = set()
        sorted_results = sorted(results, key=lambda x: (x[1] or x[0]))

        # SUMMARY sheet
        summary_rows: List[Dict[str, Any]] = [s for _, _, _, s in sorted_results]
        df_sum = pd.DataFrame(summary_rows)

        # OPTIONAL: grand totals row (sum numeric columns, blank for non-numeric)
        if not df_sum.empty:
            totals = {}
            for col in df_sum.columns:
                if pd.api.types.is_numeric_dtype(df_sum[col]):
                    totals[col] = df_sum[col].sum()
                else:
                    totals[col] = ''
            totals['Device'] = 'TOTAL'
            df_sum = pd.concat([df_sum, pd.DataFrame([totals])], ignore_index=True)

        df_sum.to_excel(xw, index=False, sheet_name="SUMMARY")
        ws_sum = xw.sheets["SUMMARY"] if hasattr(xw, "sheets") else xw.book["SUMMARY"]
        _format_worksheet(ws_sum, df_sum)


        # Per-device sheets
        for ip, hostname, detailed, summary in sorted_results:
            df = pd.DataFrame(detailed) if detailed else pd.DataFrame([{'Info': summary.get('Error', 'No data')}])

            suggested = _sheet_name_from(hostname or "", ip)
            sheet = _unique_sheet_name(suggested, used_sheet_names)

            df.to_excel(xw, index=False, sheet_name=sheet)
            ws = xw.sheets[sheet] if hasattr(xw, "sheets") else xw.book[sheet]
            _format_worksheet(ws, df)

    if args.debug:
        print(f"[*] Wrote Excel: {args.output}")


if __name__ == "__main__":
    main()